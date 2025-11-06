import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    csv_path = Path(csv_path)
    xlsx_path = Path(xlsx_path)

    if not csv_path.is_file():
        raise FileNotFoundError('файл не найден')
    
    with csv_path.open(encoding="utf-8", newline="") as f:
        r = csv.reader(f)
        rows = list(r)
    
    if not rows:
        raise ValueError("файл пустой")
    
    book = Workbook()
    sheet = book.active
    sheet.title = "Sheet1"

    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

    for col_index in range(1, len(rows[0]) + 1):
        col_letter = get_column_letter(col_index)
        max_length = max((len(str(sheet.cell(row=row, column=col_index).value or "")) for row in range(1, len(rows) + 1)), default=8)
        sheet.column_dimensions[col_letter].width = max(max_length, 8)

    book.save(xlsx_path)

csv_to_xlsx('data/lab05/samples/people.csv', 'data/lab05/out/people.xlsx')  